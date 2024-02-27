import logging

import openpyxl
import fnmatch
import os

from openpyxl.comments import Comment
from openpyxl.worksheet.worksheet import Worksheet


def enter_data_on_the_sheet(table: list, headers: dict, sheet: Worksheet, com_add: bool = True, c1: int = 0) -> None:
    """
        **Функция `enter_data_on_the_sheet`**

        Эта функция заполняет лист Excel данными из переданной таблицы, используя заголовки,
        определенные в словаре `headers`. Если в данных есть поле 'Расхождения',
        то создаются комментарии и добавляются к соответствующим ячейкам.

        **Параметры**
        - `table` (список): Таблица данных для заполнения в формате списка словарей.
        - `headers` (словарь): Словарь заголовков, где ключи - это названия полей, а значения - описания.
        - `sheet` (объект листа Excel): Лист Excel, на который нужно внести данные.
        - `com_add` (булиан): Опредиляет, нужно ли добовлять комментарии.

        **Возвращаемое значение**
        - `None`: Функция не возвращает значения, но изменяет переданный лист Excel.

        **Пример использования**
            data_table = [
                {"Name": "John", "Age": 30, "Расхождения": [{"Преф": "Diff", "Рез": 5, "Тип": "Age"}]},
                {"Name": "Alice", "Age": 25},
                # ...
            ]
            headers_info = {"Name": "Имя пользователя", "Age": "Возраст"}
            enter_data_on_the_sheet(data_table, headers_info, excel_sheet)
    """
    if not isinstance(table, list):
        return None

    if c1 != 0:
        c1 -= 1

    table.insert(0, headers)  # Вставляем заголовки в начало таблицы
    for i, row_data in enumerate(table, start=1):
        for j, key in enumerate(headers, start=1):
            sheet.cell(row=i, column=c1 + j, value=str(row_data.get(key, '')))
        if com_add and 'Расхождения' in row_data:  # Добавляем комментарии на лист если они есть
            for com in row_data['Расхождения']:
                comment = Comment(com['Преф'] + str(com['Рез']), 'Я', 100, 350)
                coll_num = get_number_key_in_dict(headers, com['Тип'])
                if coll_num:
                    sheet.cell(row=i, column=c1 + get_number_key_in_dict(headers, com['Тип'])).comment = comment


def get_list_in_excel(file_name: str, sheet_name: str, create_workbook: bool = False, one_sheet: bool = False) -> tuple:
    """
        **Функция `get_list_in_excel`**

        Эта функция открывает файл Excel и возвращает рабочую книгу и лист. Если файл не существует, и установлен флаг
            `create_workbook`, то создается новая книга.

        **Параметры**
        - `file_name` (строка): Имя файла Excel.
        - `sheet_name` (строка): Имя листа Excel.
        - `create_workbook` (булево): Флаг для создания новой книги, если файл не существует.
            По умолчанию установлен в `True`.

        **Возвращаемое значение**
        - Кортеж из двух элементов:
          - `wb` (рабочая книга): Объект рабочей книги Excel.
          - `sh` (лист Excel): Объект листа Excel.

        **Пример использования**
        workbook, sheet = get_list_in_excel("example.xlsx", "Sheet1")
    """

    if create_workbook or not os.path.exists(file_name):
        wb = openpyxl.Workbook()
    else:
        wb = openpyxl.load_workbook(file_name)

    if one_sheet:
        for sh in wb:
            if sh.title == sheet_name:
                return wb, sh
        sh = wb.create_sheet(sheet_name)
        return wb, sh

    if len(wb.worksheets) == 1:
        sh = wb.active
        if sh.title == 'Sheet':
            sh.title = sheet_name
        else:
            sh = wb.create_sheet(sheet_name)
        return wb, sh

    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    sh = wb.create_sheet(sheet_name)

    return wb, sh


def get_number_key_in_dict(dict, key):
    for i, item in enumerate(dict):
        if item == key:
            return i + 1
    return 0


def find_sheet(workbook, mask_name):
    sheets = []
    if workbook and mask_name:
        for sheet in workbook:
            if fnmatch.fnmatch(sheet.title, mask_name):
                sheets.append(sheet.title)
    return sheets


def get_data_excel(sheet):
    all_data = []
    if sheet:
        for row in sheet.iter_rows():
            filtered_row = []
            data = False
            for cell in row:
                text = cell.value
                if text is None:
                    text = ''
                filtered_row.append(str(text).strip())
                if cell.value is not None and cell.value != '':
                    data = True
            if data:
                all_data.append(filtered_row)
    return all_data


def find_headers(table, headers):
    result_headers = {}
    if table and headers:
        numbe_row = []
        for key, items in headers.items():
            if not type(items) is list:
                items = [items]
            result_headers[key] = []
            for item in items:
                if not str(item).isdigit():
                    for j, row in enumerate(table):
                        for i, cell in enumerate(row):
                            if fnmatch.fnmatch(str(cell).lower(), item.lower()):
                                result_headers[key] = i
                                numbe_row.append(j)
                                break
                        if str(result_headers[key]).isdigit():
                            break
                if str(result_headers[key]).isdigit():
                    break
        if numbe_row:
            numbe_row = min(numbe_row)
        else:
            numbe_row = 0
    return result_headers, numbe_row


def get_data_by_headers(table, headers):
    date = []
    if table and headers:
        normal_headers, row_result = find_headers(table, headers)
        for key, value in normal_headers.items():
            if isinstance(value, list):
                logging.warning(f'Не найдены заголовки - {key}!')
        if any(isinstance(value, list) and not value for value in normal_headers.values()):
            return date

        if normal_headers:
            for row in table[row_result:]:
                row_data = {}
                for key, item in normal_headers.items():
                    if isinstance(item, int) and len(row) > item:
                        row_data[key] = row[item]
                if row_data:
                    date.append(row_data)
    return date


def get_data_from_sheet(path, headers, sheet=0):
    data = []
    if path and headers:
        workbook = openpyxl.load_workbook(path, data_only=True)  # Читаем Excel

        # Находим лист в книге
        if str(sheet).isdigit():
            sh = workbook.worksheets[sheet]
        else:
            sheet_name = find_sheet(workbook, sheet)
            if not sheet_name:
                return data
            sh = sheet_name[0]

        all_data = get_data_excel(sh)  # Получаем данные с листа
        data = get_data_by_headers(all_data, headers)  # Оставить только указанные заголовки

        workbook.close()  # Закрыть книгу
    return data
