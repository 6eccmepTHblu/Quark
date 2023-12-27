from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.worksheet.worksheet import Worksheet


def conditional_formatting_color(sheet: Worksheet, color: str, formul: str, range: str) -> None:
    """
    **Функция `conditional_formatting_color`**
    Эта функция добавляет условное форматирование в лист Excel для указанного диапазона ячеек,
    используя указанный цвет и формулу.

    **Параметры**
    - `sheet` (объект листа Excel): Лист Excel, к которому нужно добавить условное форматирование.
    - `color` (строка): Цвет для условного форматирования в формате RGB (#RRGGBB) или предопределенное имя цвета.
    - `formul` (строка): Формула для условного форматирования.
    - `range` (строка): Диапазон ячеек, к которым применяется условное форматирование.

    **Возвращаемое значение**
    - `None`: Функция не возвращает значения, но изменяет переданный лист Excel, добавляя условное форматирование.

    **Пример использования**
    conditional_formatting_color(excel_sheet, "FF0000", "A1>10", "A1:A10")
    """
    if sheet and color and formul and range:
        diff_style = DifferentialStyle(fill=PatternFill(start_color=color, end_color=color, fill_type="solid"))
        rule = Rule(type="expression", formula=[formul], dxf=diff_style)
        sheet.conditional_formatting.add(range, rule)

def conditional_formatting_font(sheet: Worksheet, formul: str, range: str) -> None:
    """
    **Функция `conditional_formatting_font`**
    Эта функция добавляет условное форматирование для шрифта в лист Excel для указанного диапазона ячеек,
    используя заданную формулу.

    **Параметры**
    - `sheet` (объект листа Excel): Лист Excel, к которому нужно добавить условное форматирование для шрифта.
    - `formul` (строка): Формула для условного форматирования.
    - `range` (строка): Диапазон ячеек, к которым применяется условное форматирование для шрифта.

    **Возвращаемое значение**
    - `None`: Функция не возвращает значения, но изменяет переданный лист Excel,
                добавляя условное форматирование для шрифта.

    **Пример использования**
    conditional_formatting_font(excel_sheet, "A1>10", "A1:A10")
    """
    if sheet and formul and range:
        diff_style = DifferentialStyle(font=Font(bold=True))
        rule = Rule(type="expression", formula=[formul], dxf=diff_style)
        sheet.conditional_formatting.add(range, rule)

def get_column_by_key(headers: list, key: str|list) -> None:
    """
    **Функция `get_column_by_key`**
    Эта функция возвращает диапазон столбцов в формате Excel (например, "A1:B1") для заданных ключей из заголовков.

    **Параметры**
    - `headers` (список): Список заголовков, представляющих собой названия столбцов.
    - `key` (строка или список строк): Ключ или ключи для определения столбца или диапазона столбцов.

    **Возвращаемое значение**
    - Строка: Диапазон столбцов в формате Excel.

    **Пример использования**
    columns_range = get_column_by_key(["Name", "Age", "Gender"], "Age")
    """
    if headers and key:
        if not isinstance(key, list):
            key = [key]

        range = []
        for elem in key:
            for i, item in enumerate(headers):
                if item == elem:
                    range.append(i + 1)
                    break
        if range:
            if len(range) == 1:
                return '$' + get_column_letter(range[0]) + '1'
            else:
                return get_column_letter(range[0]) + '1:' + get_column_letter(range[1]) + '1048576'
    return ''

def conditional_formatting(uf:dict, sh: Worksheet) -> None:
    """
    **Функция `conditional_formatting`**
    Эта функция применяет условное форматирование к листу Excel на основе предоставленных правил форматирования.

    **Параметры**
    - `uf` (словарь): Словарь с правилами условного форматирования, где ключи - это формулы,
            а значения - это списки правил для каждой формулы.
    - `sh` (объект листа Excel): Лист Excel, к которому нужно применить условное форматирование.

    **Возвращаемое значение**
    - `None`: Функция не возвращает значения, но изменяет переданный лист Excel, применяя условное форматирование.

    **Пример использования**
    formatting_rules = {
        "A1>10": [{"Диапазон": "A1:A10", "Цвет": "FF0000"}],
        "B1<5": [{"Диапазон": "B1:B10", "Цвет": "00FF00"}]
    }

    conditional_formatting(formatting_rules, excel_sheet)
    """
    if uf and sh:
        for key, item in uf.items():
            for elem in item:
                if not isinstance(elem['Диапазон'], list):
                    elem['Диапазон'] = [elem['Диапазон']]
                for range in elem['Диапазон']:
                    conditional_formatting_color(sh, elem['Цвет'], key, range)
