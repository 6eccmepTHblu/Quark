import csv
import glob
import logging
import openpyxl
import os

from datetime import datetime
from fnmatch import fnmatch
from consts.files import FILES
from def_folder.normalization import (normalisation_par_type_de_fichier,
                                      normalization_of_data_by_headers)
from openpyxl.comments import Comment
from openpyxl.worksheet.worksheet import Worksheet
from def_folder.data_normalization import transliteration_ru_en
from def_folder.csv import get_data_in_csv

# Глобальные переменные для хранения данных нормирования заголовков
loaded_data = []


def load_data():
    global loaded_data

    path = os.path.abspath(__file__).replace(os.path.basename(__file__), '')
    path_csv = os.path.join(path, 'normalization', '_ЗАГОЛОВКИ.csv')
    if os.path.exists(path_csv):
        loaded_data = get_data_in_csv([path_csv], name=False, take_strio=False)


def find_file(path: str, templates: list[str]) -> str:
    """
        **Функция `find_file`**
        Эта функция принимает путь к директории и список шаблонов файлов и возвращает путь к первому найденному файлу,
        соответствующему одному из шаблонов.

        **Параметры**
        - `path` (строка): Путь к директории, в которой нужно искать файлы.
        - `templates` (список): Список шаблонов файлов, которые нужно найти в директории.

        **Возвращаемое значение**
        - Путь к файлу (строка): Путь к первому найденному файлу, соответствующему одному из шаблонов.
            Если не найдено подходящих файлов или параметры функции не корректны, функция вернет `None`.
    """
    if not path or not templates or not os.path.exists(path):
        return ''

    for file_name in os.listdir(path):
        for mask in templates:
            if fnmatch(file_name, mask):
                return os.path.join(path, file_name)

    return ''


def get_list_files(path: str, templates: list[str]) -> list[str]:
    """
        **Функция `get_list_files`**
        Эта функция принимает путь к директории и список шаблонов файлов и возвращает список файлов,
        соответствующих этим шаблонам.

        **Параметры**
        - `path` (строка): Путь к директории, в которой нужно искать файлы.
        - `templates` (список): Список шаблонов файлов, которые нужно найти в директории.

        **Возвращаемое значение**
        - Список файлов (список): Список файлов, соответствующих шаблонам.
            Если не найдено подходящих файлов или параметры функции не корректны, функция вернет пустой список.
    """
    list_files = []
    if not path or not templates or not os.path.exists(path):
        return list_files

    for mask in templates:
        files = glob.glob(os.path.join(path, mask))
        list_files.extend(files)

    return list_files


def separate_collumns_of_headers(all_data: list[list], headers: dict, group: str, path: bool = 'True', file='') -> list[
    list]:
    """
    **Функция `separate_collumns_of_headers`**
    Эта функция создает список словарей, представляющих собой выбранные столбцы данных из каждой строки `all_data`
    в соответствии с заданными заголовками `headers`. Каждый словарь также содержит дополнительное поле `'ТипФайла'`,
    указывающее на принадлежность к определенной группе.

    **Параметры**
    - `all_data` (список списков): Двумерный список данных, где каждый внутренний список
        представляет собой строку данных.
    - `headers` (словарь): Словарь, где ключи - это имена столбцов,
        а значения - это индексы столбцов в каждой строке `all_data`.
    - `group` (строка): Имя группы, к которой относятся данные.
    - `path` (строка): Если установлено в `True`, добавляет в словарь дополнительное поле `'ПутьФайла'`,
        содержащее путь к файлу. По умолчанию установлено в `True`.

    **Возвращаемое значение**
    - `data` (список словарей): Список словарей, представляющих выбранные столбцы данных
        с дополнительным полем `'ТипФайла'`, если path = True.

    **Пример использования**
    all_data = [["John", 25, "Male"], ["Jane", 30, "Female"]]
    headers = {"Name": 0, "Age": 1, "Gender": 2}
    group = "People"
    result_data = separate_collumns_of_headers(all_data, headers, group, path=True)
    """
    data = []
    if not all_data or not headers:
        return data

    if path:
        headers['ПутьФайла'] = -1
    for row in all_data:
        temp_dict = {}

        for key, item in headers.items():
            if isinstance(item, list):
                item = item[0]
            try:
                if isinstance(item, str):
                    hed = item
                    item, _ = find_headers(all_data, {"name": item})
                    item = item["name"]
                    if isinstance(item, list):
                        logging.warning(f'Не найдены заголовки - "{hed}" в "{group}" - "{file}"!')
                        return []
                temp_dict[key] = row[item] if len(row) > item else ''
            except IndexError:
                temp_dict[key] = ''
        temp_dict['ТипФайла'] = group

        if temp_dict:
            data.append(temp_dict)
    return data


def crossing_tables(table_1: list[dict], table_2: list[dict], header: str) -> list[dict]:
    """
    **Функция `crossing_tables`**
    Эта функция объединяет две таблицы данных (`table_1` и `table_2`) по заданному заголовку (`header`).
    Объединение происходит при условии равенства значения в указанном столбце.

    **Параметры**
    - `table_1` (список словарей): Первая таблица данных, представленная списком словарей.
    - `table_2` (список словарей): Вторая таблица данных, представленная списком словарей.
    - `header` (строка): Название столбца, по которому происходит объединение таблиц.

    **Возвращаемое значение**
    - `table_1` (список словарей): Обновленная первая таблица данных после выполнения объединения с таблицей `table_2`.

    **Пример использования**
    table1 = [{"ID": 1, "Name": "John"}, {"ID": 2, "Name": "Jane"}]
    table2 = [{"ID": 2, "Age": 30}, {"ID": 1, "Age": 25}]
    result_table = crossing_tables(table1, table2, "ID")
    """
    if not table_1 or not table_2 or not header:
        return table_1

    for row_1 in table_1:
        for row_2 in table_2:
            if row_1.get(header, '-') == row_2.get(header, '+'):
                for key, item in row_2.items():
                    if key in row_1:
                        key += "|"
                    row_1[key] = item
                break
    return table_1


def pull_out_the_page_namber(data: list[dict]) -> list[dict]:
    """
    **Функция `pull_out_the_page_number`**
    Эта функция извлекает номер страницы из пути к файлу в каждой строке данных словаря `data`
    и добавляет его как новое поле с ключом `'СтраницаОбщая'`.

    **Параметры**
    - `data` (список словарей): Список словарей, представляющих данные, в каждом из которых есть поле `'ПутьФайла'`.

    **Возвращаемое значение**
    - `data` (список словарей): Список словарей с добавленным полем `'СтраницаОбщая'` в каждой строке данных.

    **Пример использования**
    data = [{"ПутьФайла": "folder1/file_1_10.xlsx"}, {"ПутьФайла": "folder2/file_2_25.xlsx"}]
    result_data = pull_out_the_page_number(data)
    """
    if data:
        for row in data:
            path_file = row.get('ПутьФайла', '')
            if path_file:
                page_number = path_file.split('_')[-1].split('.')[0]
                row['СтраницаОбщая'] = page_number
    return data


def collects_data_by_type(path: str, types: list, table_csv: str = '', norml=True) -> list:
    all_data = []
    if not path or not types:
        return all_data

    for type in types:
        if type not in FILES:
            logging.warning(f'Данного типа "{type}" нет в КОНСТАНТЕ!')
            continue

        # Получаем данные по заголовкам
        data_list = FILES[type]
        if not isinstance(data_list, list):
            data_list = [data_list]

        for data_dict in data_list:
            logging.info(f'Сбор данных из {data_dict["Маска"]}.')
            data_csv = get_data_from_csv(path, data_dict['Маска'], data_dict['Заголовки'], data_dict.get('Отступ', 1),
                                         type, norml)

            # Если необходимы табличные данные, то дабовляем их
            table_csv_name = ''
            if table_csv != '-' and 'Таблица' in data_dict:
                # Превращяем в лист, если строка
                if isinstance(data_dict['Таблица'], str):
                    data_dict['Таблица'] = [data_dict['Таблица']]

                # Проверка, есть ли указаныя таблица
                if table_csv == '':
                    table_csv_name = data_dict['Таблица'][0]
                elif table_csv in data_dict['Таблица']:
                    table_csv_name = table_csv

                if table_csv_name != '':
                    logging.info(f'Сбор данных из {table_csv_name}.')
                    table_type = FILES[table_csv_name]
                    data_table_csv = get_data_from_csv(path, table_type['Маска'], table_type['Заголовки'],
                                                       table_type.get('Отступ', 1), table_csv_name, norml)
                    data_table_csv = pull_out_the_page_namber(
                        data_table_csv)  # Вытаскиваем номер страницы из имени файла
                    data_csv = crossing_tables(data_table_csv, data_csv, 'СтраницаОбщая')

            # Если данные есть, вносим в общие данные
            logging.info(f'     Данные из {data_dict["Маска"]} - [{table_csv_name}] собранны - {str(len(data_csv))}')
            if data_csv:
                all_data += data_csv

    return all_data


def get_data_from_csv(path, masks, headers, indent, group, norml):
    # Получаем список файлов
    list_files = get_list_files(path, masks)

    # Собираем все данные из выбранных файлов
    data_all = []
    for file in list_files:
        total_data = get_data_in_csv(file)

        # Выбераем нужные столбцы, разбивая по словарю
        data = separate_collumns_of_headers(total_data, headers, group, file=file)
        data_all.extend(data[indent:])

    # Нормализация данных по типу
    if norml:
        data_all = normalisation_par_type_de_fichier(data_all, group)

    # Нормализация данных по заголовкам
    normal_data = normalization_of_data_by_headers(data_all, headers, group)

    # Убераем пробелы по бокам
    for row in normal_data:
        for key, value in row.items():
            if isinstance(value, str):
                row[key] = value.strip()

    return normal_data


def output_data_to_csv(path, table, name):
    if path and table and name:
        folder = 'Данные сверки'

        # Если папки нет, то создаем
        folder_path = path + '\\' + folder
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)

        # Если указанный файл есть, то создаем его
        folder_path = folder_path + '\\' + name + '.csv'
        try:
            with open(folder_path, mode='w', newline='') as csv_file:
                # Создаем объект writer
                csv_writer = csv.DictWriter(csv_file, fieldnames=table[0].keys(), delimiter=';')

                # Записываем заголовки
                csv_writer.writeheader()

                # Записываем данные из списка со словарями
                for row in table:
                    csv_writer.writerow(row)
            print('Создана записать в csv - ' + name + '.csv')
        except Exception as _:
            print('Не удалось записать в csv - ' + name + '.csv')


def checks_dates(date_1: str, date_2: str, type_check: str = '!=') -> bool:
    date_1 = types_date(date_1)
    date_2 = types_date(date_2)

    if isinstance(date_1, str) or isinstance(date_2, str):
        if type_check == '==':
            if str(date_1) == str(date_2):
                return True
            elif type_check == '!=':
                if str(date_1) != str(date_2):
                    return True
    else:
        if type_check == '==' and date_1 == date_2:
            return True
        elif type_check == '!=' and date_1 != date_2:
            return True
        elif type_check == '<' and date_1 < date_2:
            return True
        elif type_check == '>' and date_1 > date_2:
            return True
        elif type_check == '<=' and date_1 <= date_2:
            return True
        elif type_check == '>=' and date_1 >= date_2:
            return True

    return False


def types_date(date: str) -> str | datetime:
    if date == '':
        return ''

    try:
        return datetime.strptime(date, '%d.%m.%Y')
    except Exception as _:
        return date


def sums_rows_by_keys(table: list[dict], keys: list[str], keys_sum: list[str]) -> list[dict]:
    if not table or not keys and not keys_sum:
        return table

    key_name = '__ключ_слияния__'
    for row in table:
        row[key_name] = '!'.join([str(row.get(key, '')) for key in keys])

    set_keys = set(row[key_name] for row in table)

    all_data = []
    for key in set_keys:
        filter_data = [row for row in table if row[key_name] == key]
        if len(filter_data) == 1:
            all_data.extend(filter_data)
        else:
            number = -1
            for row in filter_data:
                result_sum = True
                for key_sum in keys_sum:
                    try:
                        result_sum *= isinstance(float(str(row.get(key_sum, '+'))), float)
                    except:
                        result_sum *= False
                        break
                if result_sum:
                    if number == -1:
                        all_data.extend([row])
                        number = len(all_data) - 1
                    else:
                        for key_sum in keys_sum:
                            all_data[number][key_sum] = float(all_data[number][key_sum]) + float(row[key_sum])
                            all_data[number][key_sum] = str(all_data[number][key_sum]).replace('.0', '')
                else:
                    all_data.extend([row])
    return all_data


def enter_data_on_the_sheet(table: list, headers: dict, sheet: Worksheet, com_add: bool = True, c1: int = 0) -> None:
    """
        **Функция `enter_data_on_the_sheet`**

        Эта функция заполняет лист Excel данными из переданной таблицы, используя заголовки,
        определенные в словаре `headers`. Если в данных есть поле 'Расхождения',
        то создаются комментарии и добавляются к соответствующим ячейкам.

        **Параметры**
        - `table` (список): Таблица данных для заполнения в формате списка словарей.
        - `headers` (словарь): Словарь заголовков, где ключи - это названия полей, а значения - описания.
        - `sheet` (объект листа Excel): Лист Excel, на который нужно внести данные.
        - `com_add` (булиан): Опредиляет, нужно ли добовлять комментарии.

        **Возвращаемое значение**
        - `None`: Функция не возвращает значения, но изменяет переданный лист Excel.

        **Пример использования**
            data_table = [
                {"Name": "John", "Age": 30, "Расхождения": [{"Преф": "Diff", "Рез": 5, "Тип": "Age"}]},
                {"Name": "Alice", "Age": 25},
                # ...
            ]
            headers_info = {"Name": "Имя пользователя", "Age": "Возраст"}
            enter_data_on_the_sheet(data_table, headers_info, excel_sheet)
    """
    if not isinstance(table, list):
        return None

    if c1 != 0:
        c1 -= 1

    table.insert(0, headers)  # Вставляем заголовки в начало таблицы
    for i, row_data in enumerate(table, start=1):
        for j, key in enumerate(headers, start=1):
            sheet.cell(row=i, column=c1 + j, value=str(row_data.get(key, '')))
        if com_add and 'Расхождения' in row_data:  # Добавляем комментарии на лист если они есть
            for com in row_data['Расхождения']:
                comment = Comment(com['Преф'] + str(com['Рез']), 'Я', 100, 350)
                coll_num = get_number_key_in_dict(headers, com['Тип'])
                if coll_num:
                    sheet.cell(row=i, column=c1 + get_number_key_in_dict(headers, com['Тип'])).comment = comment


def get_list_in_excel(file_name: str, sheet_name: str, create_workbook: bool = False, one_sheet: bool = False) -> tuple:
    """
        **Функция `get_list_in_excel`**

        Эта функция открывает файл Excel и возвращает рабочую книгу и лист. Если файл не существует, и установлен флаг
            `create_workbook`, то создается новая книга.

        **Параметры**
        - `file_name` (строка): Имя файла Excel.
        - `sheet_name` (строка): Имя листа Excel.
        - `create_workbook` (булево): Флаг для создания новой книги, если файл не существует.
            По умолчанию установлен в `True`.

        **Возвращаемое значение**
        - Кортеж из двух элементов:
          - `wb` (рабочая книга): Объект рабочей книги Excel.
          - `sh` (лист Excel): Объект листа Excel.

        **Пример использования**
        workbook, sheet = get_list_in_excel("example.xlsx", "Sheet1")
    """

    if create_workbook or not os.path.exists(file_name):
        wb = openpyxl.Workbook()
    else:
        wb = openpyxl.load_workbook(file_name)

    if one_sheet:
        for sh in wb:
            if sh.title == sheet_name:
                return wb, sh
        sh = wb.create_sheet(sheet_name)
        return wb, sh

    if len(wb.worksheets) == 1:
        sh = wb.active
        if sh.title == 'Sheet':
            sh.title = sheet_name
        else:
            sh = wb.create_sheet(sheet_name)
        return wb, sh

    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    sh = wb.create_sheet(sheet_name)

    return wb, sh


def get_number_key_in_dict(dict, key):
    for i, item in enumerate(dict):
        if item == key:
            return i + 1
    return 0


def find_sheet(workbook, mask_name):
    sheets = []
    if workbook and mask_name:
        for sheet in workbook:
            if fnmatch(sheet.title, mask_name):
                sheets.append(sheet.title)
    return sheets


def get_data_excel(sheet):
    all_data = []
    if sheet:
        for row in sheet.iter_rows():
            filtered_row = []
            data = False
            for cell in row:
                text = cell.value
                if text is None:
                    text = ''
                filtered_row.append(str(text).strip())
                if cell.value is not None and cell.value != '':
                    data = True
            if data:
                all_data.append(filtered_row)
    return all_data


def find_headers(table, headers):
    result_headers = {}
    if table and headers:
        numbe_row = []
        for key, items in headers.items():
            if not type(items) is list:
                items = [items]
            result_headers[key] = []
            for item in items:
                item = normalizes_headings(item)
                item = transliteration_ru_en(item, 'ru')
                if not str(item).isdigit():
                    for j, row in enumerate(table):
                        for i, cell in enumerate(row):
                            cell = normalizes_headings(cell)
                            cell = transliteration_ru_en(cell, 'ru')
                            if fnmatch(str(cell).lower(), item.lower()):
                                result_headers[key] = i
                                numbe_row.append(j)
                                break
                        if str(result_headers[key]).isdigit():
                            break
                if str(result_headers[key]).isdigit():
                    break
        if numbe_row:
            numbe_row = min(numbe_row)
        else:
            numbe_row = 0
    return result_headers, numbe_row


def get_data_by_headers(table, headers):
    date = []
    if table and headers:
        normal_headers, row_result = find_headers(table, headers)
        for key, value in normal_headers.items():
            if isinstance(value, list):
                logging.warning(f'Не найдены заголовки - {key}!')
        if any(isinstance(value, list) and not value for value in normal_headers.values()):
            return date

        if normal_headers:
            for row in table[row_result:]:
                row_data = {}
                for key, item in normal_headers.items():
                    if isinstance(item, int) and len(row) > item:
                        row_data[key] = row[item]
                if row_data:
                    date.append(row_data)
    return date


def get_data_from_sheet(path, headers, sheet=0):
    data = []
    if path and headers:
        workbook = openpyxl.load_workbook(path, data_only=True)  # Читаем Excel

        # Находим лист в книге
        if str(sheet).isdigit():
            sh = workbook.worksheets[sheet]
        else:
            sheet_name = find_sheet(workbook, sheet)
            if not sheet_name:
                return data
            sh = sheet_name[0]

        all_data = get_data_excel(sh)  # Получаем данные с листа
        data = get_data_by_headers(all_data, headers)  # Оставить только указанные заголовки

        workbook.close()  # Закрыть книгу
    return data


def normalizes_headings(header):
    if isinstance(header, str):
        for row in loaded_data:
            if len(row) == 2:
                header = header.replace(row[0], row[1])
    return header
