import logging
import openpyxl
import fnmatch
import re

from def_folder import data_collection as fun
from def_folder import data_normalization as norm

DATE_CELL = '*(дата составления акта)*'
NUMBER_CELL = [['*освидетельствования строительных конструкций*', 2, 1],
               ['*освидетельствования скрытых работ*', 1, 1]]
SHEET_REESTR = '*реестр*'
TEMPLATES = ['*АООК*.xls*',
             '*Акт освидетельствования ответственных конструкций*.xls*',
             '*АОСР*.xls*']
HEADERS = {'Пункт': '№ п/п',
           'ИД': '*документ*',
           'Номер/Дата': '*Реквизит*'}


def get_data(path, mask, type_name):
    # Найти файл
    file = fun.find_file(path, TEMPLATES)
    logging.info('Сбор данных из АООК/АОСР.')
    if file == '':
        logging.warning(f"Не найден файл АООК/АОСР - '{TEMPLATES}'")
        return None

    # Открываем книгу АООК
    wb = openpyxl.load_workbook(file)

    # Собираем реестры АООК
    reestr_aook = []
    for sh in wb:
        if fnmatch.fnmatch(sh.title.lower(), SHEET_REESTR):
            all_data = fun.get_data_excel(sh)

            # Ищем нужные столбцы
            headers, resul_rows = fun.find_headers(all_data, HEADERS)

            # Выбераем нужные столбцы, разбивая по словарю
            data = []
            for row in all_data[resul_rows:]:
                row_data = {}
                for key, item in headers.items():
                    if type(item) is list:
                        logging.warning(f'Заголовок "{HEADERS[key]}" - не был найден в АООК/АОСР!')
                        return reestr_aook
                    row_data[key] = row[item]
                data.append(row_data)

            # Фильтруем данные по АоРПИ
            for row in data:
                if fnmatch.fnmatch(str(row['ИД']).lower(), mask):
                    reestr_aook.append(row)
            pass

    # Нормирование
    for row in reestr_aook:
        row['Путь файла АООК'] = file
        if 'Номер/Дата' in row:
            # Дата
            date = row['Номер/Дата'].split('от')
            if len(date) > 1:
                date = date[1]
            else:
                date = date[0]
            row[f'{type_name} Дата из АООК'] = norm.get_date(date)

            # Номер
            nambe = row['Номер/Дата'].split('от')[0].strip().replace('№', '')
            nambe = norm.transliteration_ru_en(nambe).upper()
            nambe = delete_in_brackets(nambe)
            row[f'{type_name} Номер из АООК'] = nambe

    logging.info('      Данные из АООК/АОСР собранны - ' + str(len(reestr_aook)) + '.')
    return reestr_aook


def delete_in_brackets(text: str) -> str:
    # Используем регулярное выражение для поиска и удаления содержимого в скобках
    pattern = re.compile(r'\([^)]*\)')
    result = re.sub(pattern, '', text)
    return result


def get_date(path):
    # Найти файл
    file = fun.find_file(path, TEMPLATES)
    if file == '':
        logging.warning(f"Не найден файл АООК/АОСР - '{TEMPLATES}'")
        return None

    # Открываем книгу АООК
    wb = openpyxl.load_workbook(file)

    # Находим дату АООК
    sh = wb[wb.sheetnames[0]]
    data_aook = ''
    for row in sh.iter_rows():
        for cell in row:
            if fnmatch.fnmatch(str(cell.value), DATE_CELL):
                # Если нашли ячейку со значением "дата", то берем значение из строки выше
                data_aook = sh.cell(row=cell.row - 1, column=cell.column).value
                if data_aook is None:
                    data_aook = sh.cell(row=cell.row - 1, column=cell.column - 1).value
                break
        if data_aook:
            break

    # Нормируем дату АООК
    data_aook = norm.get_date(data_aook)
    data_aook = data_aook.replace('"', '').replace('г.', '').replace('-', '.')
    if ' ' in data_aook:
        data_aook = data_aook.split(' ')[0].strip()
    data_aook = norm.get_date(data_aook)

    logging.info(f"Дата АООК/АОСР: {data_aook}")
    return data_aook


def get_number(path):
    # Найти файл
    file = fun.find_file(path, TEMPLATES)
    if file == '':
        logging.warning(f"Не найден файл АООК/АОСР - '{TEMPLATES}'")
        return None

    # Открываем книгу АООК
    wb = openpyxl.load_workbook(file)

    # Находим номер АООК
    sh = wb[wb.sheetnames[0]]
    number_aook = ''
    for row in sh.iter_rows():
        for cell in row:
            for number in NUMBER_CELL:
                if fnmatch.fnmatch(str(cell.value), number[0]):
                    number_aook = sh.cell(row=cell.row + number[1], column=cell.column + number[2]).value
                    break
        if number_aook:
            break

    # Нормируем дату АООК
    logging.info(f"Номер АООК/АОСР: {number_aook}")

    return number_aook


if __name__ == '__main__':
    get_number('D:\Работа\Силенко Д.Т\Задача 2(Путхон - Кварк V2)\Данные 4')
