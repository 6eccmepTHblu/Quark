from openpyxl.styles import NamedStyle, Border, Side, Alignment, Font, PatternFill, Color
from openpyxl.utils import get_column_letter as collumn
from consts.colors import COLORS
from openpyxl.worksheet.worksheet import Worksheet


class RangeStyle:
    def __init__(self, sh, style_name, cell_color = COLORS['Зелёный Оч.светлый'], wrap=True, size = 11,
                 bold = False):
        self.sh = sh
        self.style_name = style_name
        self.style = NamedStyle(name=self.style_name + '_style')
        self.cell_color = cell_color
        self.wrap = wrap
        self.size = size
        self.bold = bold
        self.setup_style()

    def setup_style(self):
        bd = Side(border_style='thin', color='000000')
        self.style.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        self.style.number_format = "@"
        self.style.alignment = Alignment(wrap_text=self.wrap, horizontal='center', vertical='center')
        self.style.font = Font(bold=self.bold, size=self.size)
        self.style.fill = PatternFill(fill_type='solid', fgColor=self.cell_color)

    def apply_style(self, r1 = 1, c1 = 1, r2 = 1, c2 = 1):
        for row in self.sh[f'{collumn(c1)}{r1}:{collumn(c2)}{r2}']:
            for cell in row:
                cell.style = self.style

    def remove_style(self):
        for row in self.sh.iter_rows():
            for cell in row:
                # Проверяем, есть ли стиль у ячейки
                if cell.has_style:
                    # Проверяем, соответствует ли стиль тому, который мы хотим удалить
                    if cell.style == self.style_name:
                        # Устанавливаем для ячейки стиль по умолчанию
                        cell.style = NamedStyle()


class SheetStyle:
    def __init__(self, sh, column_width = None, freeze_panes = None, zoom_scale = 100, color = None,
                 grid_lines = False):
        self.sh = sh
        self.column_width = column_width
        self.freeze_panes = freeze_panes
        self.zoom_scale = zoom_scale
        self.color = color
        self.grid_lines = grid_lines
        self.apply_style()

    def apply_style(self):
        self.sh.sheet_view.showGridLines = self.grid_lines  # Вид - убрать границы ячеек
        self.sh.sheet_view.zoomScale = self.zoom_scale  # Зум видемости листа

        # Ширина колонок
        if not self.column_width is None:
            column_width(self.sh, self.column_width)  # Ширина колонок

        # Закрепить облость видемости
        if type(self.freeze_panes) == list:
            self.sh.freeze_panes = f'{collumn(self.freeze_panes[0])}{self.freeze_panes[1]}'

        # Цвет ярлыка листа
        if not self.color is None:
            self.sh.sheet_properties.tabColor = Color(rgb=self.color)


def add_filter(sh: Worksheet, row1:int = 1, col1:int = 1, row2:int = 1, col2:int = 1) -> None:
    """
    **Функция `add_filter`**
    Эта функция добавляет автоматический фильтр к листу Excel, используя указанный диапазон ячеек.

    **Параметры**
    - `sh` (объект листа Excel): Лист Excel, к которому нужно добавить фильтр.
    - `row1` (целое число): Начальная строка фильтра. По умолчанию установлено в 1.
    - `col1` (целое число): Начальный столбец фильтра. По умолчанию установлено в 1.
    - `row2` (целое число): Конечная строка фильтра. По умолчанию установлено в 1.
    - `col2` (целое число): Конечный столбец фильтра. По умолчанию установлено в 1.

    **Возвращаемое значение**
    - `None`: Функция не возвращает значения, но изменяет переданный лист Excel, добавляя автоматический фильтр.

    **Пример использования**
    add_filter(excel_sheet, 1, 1, 10, 5)
    """
    sh.auto_filter.ref = f'{collumn(col1)}{row1}:{collumn(col2)}{row2}'

def column_width(sh: Worksheet, width:int = 0) -> None:
    """
    **Функция `column_width`**
    Эта функция устанавливает ширину столбцов в листе Excel в зависимости от содержимого ячеек.

    **Параметры**
    - `sh` (объект листа Excel): Лист Excel, для которого нужно установить ширину столбцов.
    - `width` (целое число): Максимальная ширина столбца. Если установлено значение,
                             то ширина столбцов будет установлена в минимум между фактической шириной содержимого
                             и указанной шириной.

    **Возвращаемое значение**
    - `None`: Функция не возвращает значения, но изменяет переданный лист Excel, устанавливая ширину столбцов.

    **Пример использования**
    column_width(excel_sheet, 15)
    """
    if sh:
        for column in sh.columns:
            max_length = 0
            column_letter = collumn(column[0].column)
            for cell in column:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value)) + 7
            if width:
                if max_length > width:
                    max_length = width
            sh.column_dimensions[column_letter].width = max_length