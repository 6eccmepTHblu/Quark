import logging
import os
import xlwings
import ctypes
import openpyxl as xl

# PATH_FILE = "G:\DiskE\IDP\Quark\doc\Акт проверки ИД Усть-Луга.xlsm"
PATH_FILE = "D:\Работа\Силенко Д.Т\Задача 2(Путхон - Кварк V2)\Проект\doc\Акт проверки ИД Усть-Луга.xlsm"
MACROS_NAME = 'копирование_листа'


def create_act_checking(path_act: str, number_aook: str, table: list[dict]):
    # Добавляем лист Акта проверки
    vba_book = xlwings.Book(PATH_FILE)
    vba_macro = vba_book.macro(MACROS_NAME)(
        path_act,
        number_aook,
        get_display_name(),
        len(table)
    )
    vba_book.close()

    # Заполняем таблицу данными
    wb = xl.load_workbook(path_act)
    sh = wb['Акт проверки']

    for i, elem in enumerate(table, vba_macro):
        sh.cell(row=i, column=2).value = i - vba_macro + 1
        sh.cell(row=i, column=3).value = elem

    wb.save(path_act)
    logging.info(f"Создан Акт проверки. Всего замечаний - {len(table)}.")


def create_reports(table: dict):
    data_filter = []
    for key, value in table.items():
        error_num = 0
        if value:
            for row in value:
                if row.get('Акт проверки', []) and isinstance(row.get('Акт проверки', []), list):
                    error_num += 1
                    data_filter.append(row['Акт проверки'][0])
            logging.info(f"'{key}' найдено замечаний - {error_num}.")

    return data_filter


def abbreviate_full_name(full_name):
    name_parts = full_name.split()

    if len(name_parts) < 3:
        return full_name

    last_name = name_parts[0]
    first_name = name_parts[1]
    middle_name = name_parts[2]

    abbreviated_first_name = f"{first_name[0]}."
    abbreviated_middle_name = f"{middle_name[0]}."

    return f"{last_name} {abbreviated_first_name} {abbreviated_middle_name}"


def get_display_name():
    GetUserNameEx = ctypes.windll.secur32.GetUserNameExW
    NameDisplay = 3

    size = ctypes.pointer(ctypes.c_ulong(0))
    GetUserNameEx(NameDisplay, None, size)

    nameBuffer = ctypes.create_unicode_buffer(size.contents.value)
    GetUserNameEx(NameDisplay, nameBuffer, size)

    return abbreviate_full_name(nameBuffer.value)
